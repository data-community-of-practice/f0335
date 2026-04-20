"""
Unique Author Extractor
========================
Reads the output of crossref_author_fetch.py (an Excel file with a
'Crossref_Authors' column) and produces a new Excel file listing each
unique author and the DOIs they are associated with.

Handles name normalisation:
  - "John J. Smith" = "John J Smith" = "John Smith"
    = "JJ Smith" = "Smith John" = "Smith, Jihn J."

Algorithm:
  1. Strip periods, commas, extra whitespace and lowercase everything.
  2. Detect "Family, Given" format and flip to "Given Family".
  3. Try both orderings (first-last and last-first) when matching.
  4. Two names match if they share the same family name AND one's given-name
     tokens are compatible (initials match full names by first char).
  5. When merging, keep the LONGEST (most complete) name form as canonical.

Usage:
  python extract_unique_authors.py <input.xlsx> [output.xlsx]

  From Spyder:
  !python "path/to/extract_unique_authors.py" "path/to/input.xlsx"
"""

import sys
import re
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


def clean_name(name):
    name = name.strip()
    name = re.sub(r'\s+', ' ', name)
    return name


def normalise_tokens(name):
    """
    Given a raw name string, return a list of possible (family, givens_set)
    interpretations. We try:
      1. Comma format: "Smith, John J." -> family=smith, givens={smith, j}
      2. Standard order: "John J. Smith" -> family=smith, givens={john, j}
      3. Reversed order: "Smith John" -> family=smith, givens={john}
    We return all plausible interpretations so matching can try each.
    """
    # Remove periods and normalise
    name_clean = name.replace('.', '').replace('-', ' ').strip()
    name_clean = re.sub(r'\s+', ' ', name_clean)
    tokens = name_clean.lower().split()

    if len(tokens) == 0:
        return []
    if len(tokens) == 1:
        return [(tokens[0], set())]

    results = []

    # If comma format, that's definitive
    if ',' in name:
        parts = name.split(',', 1)
        family = parts[0].strip().replace('.', '').replace('-', ' ').lower()
        family = re.sub(r'\s+', ' ', family).strip()
        given_tokens = set(parts[1].strip().replace('.', '').replace('-', ' ').lower().split())
        results.append((family, given_tokens))
        return results

    # Standard: last token is family
    results.append((tokens[-1], set(tokens[:-1])))

    # Reversed: first token is family (only if first token looks like it could be a family name)
    # We add this as an alternative interpretation
    if len(tokens) >= 2:
        results.append((tokens[0], set(tokens[1:])))

    return results


def initials_expand(token):
    """
    If token is multiple uppercase initials concatenated (e.g., 'jj'),
    split into individual initials ['j', 'j'].
    """
    if len(token) <= 1:
        return [token]
    if all(c.isalpha() for c in token) and len(token) <= 3:
        # Could be concatenated initials like "jj" or "sj"
        return list(token)
    return [token]


def givens_compatible(givens_a, givens_b):
    """
    Check if two sets of given-name tokens are compatible.
    Handles: full names, single initials, and concatenated initials (e.g., 'jj').
    """
    # Expand concatenated initials
    expanded_a = set()
    for g in givens_a:
        expanded_a.update(initials_expand(g))
    expanded_b = set()
    for g in givens_b:
        expanded_b.update(initials_expand(g))

    smaller, larger = (expanded_a, expanded_b) if len(expanded_a) <= len(expanded_b) else (expanded_b, expanded_a)

    if len(smaller) == 0:
        return True

    for s_tok in smaller:
        matched = False
        for l_tok in larger:
            if s_tok == l_tok:
                matched = True
                break
            if len(s_tok) == 1 and l_tok.startswith(s_tok):
                matched = True
                break
            if len(l_tok) == 1 and s_tok.startswith(l_tok):
                matched = True
                break
        if not matched:
            return False
    return True


class AuthorCluster:
    """Represents a group of name variants that refer to the same person."""

    def __init__(self, canonical_name, family, givens):
        self.canonical_name = canonical_name
        self.family = family
        self.givens = givens
        self.all_variants = {canonical_name}
        self.dois = set()

    def matches(self, interpretations):
        """Check if any interpretation of the new name matches this cluster."""
        for family, givens in interpretations:
            if self.family == family and givens_compatible(self.givens, givens):
                return True
        return False

    def merge(self, name, interpretations, dois):
        self.all_variants.add(name)
        self.dois.update(dois)
        # Keep the longest name as canonical
        if len(name) > len(self.canonical_name):
            self.canonical_name = name
        # Find the matching interpretation and expand givens
        for family, givens in interpretations:
            if self.family == family:
                for g in givens:
                    for expanded in initials_expand(g):
                        replaced = False
                        for existing in list(self.givens):
                            if len(existing) == 1 and len(expanded) > 1 and expanded.startswith(existing):
                                self.givens.discard(existing)
                                self.givens.add(expanded)
                                replaced = True
                                break
                        if not replaced:
                            already = any(
                                (len(expanded) == 1 and ex.startswith(expanded)) or expanded == ex
                                for ex in self.givens
                            )
                            if not already:
                                self.givens.add(expanded)
                break


def extract_authors(input_file, output_file=None):
    input_path = Path(input_file)
    if not input_path.is_absolute():
        script_dir = Path(__file__).resolve().parent
        if not input_path.exists():
            fallback = script_dir / input_path
            if fallback.exists():
                input_path = fallback
    input_path = input_path.resolve()

    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}")
        sys.exit(1)

    if output_file is None:
        output_path = input_path.parent / f"{input_path.stem}_unique_authors.xlsx"
    else:
        output_path = Path(output_file).resolve()

    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")

    # --- Read input ---
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = {cell.value: cell.column for cell in ws[1]}
    authors_col = headers.get("Crossref_Authors")
    doi_col = headers.get("Publication_DOI")

    if authors_col is None:
        print("ERROR: Could not find 'Crossref_Authors' column.")
        sys.exit(1)
    if doi_col is None:
        print("ERROR: Could not find 'Publication_DOI' column.")
        sys.exit(1)

    # --- Collect raw author -> DOI mappings ---
    raw_author_dois = defaultdict(set)
    skip_prefixes = ("[DOI not found]", "[No authors listed]", "[Error")

    for row in range(2, ws.max_row + 1):
        authors_val = ws.cell(row=row, column=authors_col).value
        doi_val = ws.cell(row=row, column=doi_col).value

        if not authors_val or not doi_val:
            continue
        authors_str = str(authors_val)
        if any(authors_str.startswith(p) for p in skip_prefixes):
            continue

        doi = str(doi_val).strip()
        for author in authors_str.split(';'):
            author = clean_name(author)
            if author:
                raw_author_dois[author].add(doi)

    print(f"Found {len(raw_author_dois)} raw author name variants across all DOIs.")

    # --- Cluster authors by normalised name ---
    clusters = []

    for raw_name, dois in raw_author_dois.items():
        interpretations = normalise_tokens(raw_name)
        if not interpretations:
            continue

        matched = False
        for cluster in clusters:
            if cluster.matches(interpretations):
                cluster.merge(raw_name, interpretations, dois)
                matched = True
                break

        if not matched:
            family, givens = interpretations[0]  # primary interpretation
            c = AuthorCluster(raw_name, family, givens)
            c.dois.update(dois)
            clusters.append(c)

    # --- Second pass: merge clusters that now overlap after expansion ---
    merged = True
    while merged:
        merged = False
        new_clusters = []
        consumed = set()
        for i, c1 in enumerate(clusters):
            if i in consumed:
                continue
            for j in range(i + 1, len(clusters)):
                if j in consumed:
                    continue
                c2 = clusters[j]
                if c1.family == c2.family and givens_compatible(c1.givens, c2.givens):
                    # Merge c2 into c1
                    for v in c2.all_variants:
                        interps = normalise_tokens(v)
                        c1.merge(v, interps, c2.dois)
                    consumed.add(j)
                    merged = True
            new_clusters.append(c1)
        clusters = new_clusters

    print(f"Resolved to {len(clusters)} unique authors after normalisation.")

    # --- Sort clusters alphabetically by canonical name ---
    clusters.sort(key=lambda c: c.canonical_name.lower())

    # --- Write output Excel ---
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Unique Authors"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="left", vertical="center")

    col_headers = ["Author_Name", "Name_Variants", "DOI_Count", "DOIs"]
    for col_idx, header in enumerate(col_headers, 1):
        cell = out_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    data_font = Font(name="Arial")
    for row_idx, cluster in enumerate(clusters, 2):
        variants = "; ".join(sorted(cluster.all_variants - {cluster.canonical_name}))
        dois_str = "; ".join(sorted(cluster.dois))

        out_ws.cell(row=row_idx, column=1, value=cluster.canonical_name).font = data_font
        out_ws.cell(row=row_idx, column=2, value=variants if variants else "").font = data_font
        out_ws.cell(row=row_idx, column=3, value=len(cluster.dois)).font = data_font
        out_ws.cell(row=row_idx, column=4, value=dois_str).font = data_font

    out_ws.column_dimensions['A'].width = 35
    out_ws.column_dimensions['B'].width = 50
    out_ws.column_dimensions['C'].width = 12
    out_ws.column_dimensions['D'].width = 80
    out_ws.freeze_panes = "A2"
    out_ws.auto_filter.ref = f"A1:D{out_ws.max_row}"

    out_wb.save(output_path)
    print(f"Done. {len(clusters)} unique authors saved to: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_unique_authors.py <input.xlsx> [output.xlsx]")
        sys.exit(1)
    in_arg = sys.argv[1]
    out_arg = sys.argv[2] if len(sys.argv) > 2 else None
    extract_authors(in_arg, out_arg)
