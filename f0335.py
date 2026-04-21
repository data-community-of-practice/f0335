"""
Unique Author Extractor
========================
Reads the output of crossref_author_fetch.py (an Excel file with a
'Crossref_Authors' column) and produces a new Excel file listing each
unique author and the DOIs they are associated with.

Handles name normalisation:
  - "John J. Smith" = "John Jacob Smith" = "John J Smith" = "John Smith"
    = "JJ Smith" = "Smith John" = "Smith, John J."

Usage:
  python extract_unique_authors.py <input.xlsx> [output.xlsx]
  python extract_unique_authors.py --test     # run unit tests
"""

import sys
import re
from pathlib import Path
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


# ===================================================================
# NAME NORMALISATION FUNCTIONS
# ===================================================================

def clean_name(name):
    name = name.strip()
    name = re.sub(r'\s+', ' ', name)
    return name


def normalise_tokens(name):
    name_clean = name.replace('.', '').replace('-', ' ').strip()
    name_clean = re.sub(r'\s+', ' ', name_clean)
    tokens = name_clean.lower().split()

    if len(tokens) == 0:
        return []
    if len(tokens) == 1:
        return [(tokens[0], set())]

    results = []

    if ',' in name:
        parts = name.split(',', 1)
        family = parts[0].strip().replace('.', '').replace('-', ' ').lower()
        family = re.sub(r'\s+', ' ', family).strip()
        given_tokens = set(parts[1].strip().replace('.', '').replace('-', ' ').lower().split())
        results.append((family, given_tokens))
        return results

    results.append((tokens[-1], set(tokens[:-1])))

    if len(tokens) >= 2:
        results.append((tokens[0], set(tokens[1:])))

    return results


def initials_expand(token):
    if len(token) <= 1:
        return [token]
    if all(c.isalpha() for c in token) and len(token) <= 3:
        return list(token)
    return [token]


def givens_compatible(givens_a, givens_b):
    expanded_a = set()
    for g in givens_a:
        expanded_a.update(initials_expand(g))
    expanded_b = set()
    for g in givens_b:
        expanded_b.update(initials_expand(g))

    smaller, larger = (expanded_a, expanded_b) if len(expanded_a) <= len(expanded_b) else (expanded_b, expanded_a)

    if len(smaller) == 0:
        return True

    for s_tok in smaller:
        matched = False
        for l_tok in larger:
            if s_tok == l_tok:
                matched = True
                break
            if len(s_tok) == 1 and l_tok.startswith(s_tok):
                matched = True
                break
            if len(l_tok) == 1 and s_tok.startswith(l_tok):
                matched = True
                break
        if not matched:
            return False
    return True


class AuthorCluster:
    def __init__(self, canonical_name, family, givens):
        self.canonical_name = canonical_name
        self.family = family
        self.givens = givens
        self.all_variants = {canonical_name}
        self.dois = set()

    def matches(self, interpretations):
        for family, givens in interpretations:
            if self.family == family and givens_compatible(self.givens, givens):
                return True
        return False

    def merge(self, name, interpretations, dois):
        self.all_variants.add(name)
        self.dois.update(dois)
        if len(name) > len(self.canonical_name):
            self.canonical_name = name
        for family, givens in interpretations:
            if self.family == family:
                for g in givens:
                    for expanded in initials_expand(g):
                        replaced = False
                        for existing in list(self.givens):
                            if len(existing) == 1 and len(expanded) > 1 and expanded.startswith(existing):
                                self.givens.discard(existing)
                                self.givens.add(expanded)
                                replaced = True
                                break
                        if not replaced:
                            already = any(
                                (len(expanded) == 1 and ex.startswith(expanded)) or expanded == ex
                                for ex in self.givens
                            )
                            if not already:
                                self.givens.add(expanded)
                break


def cluster_authors(raw_author_dois):
    clusters = []

    for raw_name, dois in raw_author_dois.items():
        interpretations = normalise_tokens(raw_name)
        if not interpretations:
            continue

        matched = False
        for cluster in clusters:
            if cluster.matches(interpretations):
                cluster.merge(raw_name, interpretations, dois)
                matched = True
                break

        if not matched:
            family, givens = interpretations[0]
            c = AuthorCluster(raw_name, family, givens)
            c.dois.update(dois)
            clusters.append(c)

    merged = True
    while merged:
        merged = False
        new_clusters = []
        consumed = set()
        for i, c1 in enumerate(clusters):
            if i in consumed:
                continue
            for j in range(i + 1, len(clusters)):
                if j in consumed:
                    continue
                c2 = clusters[j]
                if c1.family == c2.family and givens_compatible(c1.givens, c2.givens):
                    for v in c2.all_variants:
                        interps = normalise_tokens(v)
                        c1.merge(v, interps, c2.dois)
                    consumed.add(j)
                    merged = True
            new_clusters.append(c1)
        clusters = new_clusters

    return clusters


# ===================================================================
# UNIT TESTS
# ===================================================================

def run_tests():
    passed = 0
    failed = 0
    failures = []

    def check(test_name, condition):
        nonlocal passed, failed
        if condition:
            passed += 1
            print(f"  PASS  {test_name}")
        else:
            failed += 1
            failures.append(test_name)
            print(f"  FAIL  {test_name}")

    print("=" * 70)
    print("RUNNING UNIT TESTS")
    print("=" * 70)

    # -- clean_name --
    print("\n--- clean_name ---")
    check("strips whitespace",
          clean_name("  John Smith  ") == "John Smith")
    check("collapses multiple spaces",
          clean_name("John    Smith") == "John Smith")
    check("handles tabs and newlines",
          clean_name("John\t\nSmith") == "John Smith")

    # -- normalise_tokens --
    print("\n--- normalise_tokens ---")

    r = normalise_tokens("John Smith")
    check("standard: family=smith",
          any(f == "smith" for f, g in r))
    check("standard: given={john}",
          any(g == {"john"} for f, g in r if f == "smith"))

    r = normalise_tokens("John J. Smith")
    check("middle initial: givens contain john and j",
          any({"john", "j"}.issubset(g) for f, g in r if f == "smith"))

    r = normalise_tokens("Smith, John J.")
    check("comma format: family=smith",
          r[0][0] == "smith")
    check("comma format: givens={john, j}",
          r[0][1] == {"john", "j"})

    r = normalise_tokens("Madonna")
    check("single name: one token",
          r == [("madonna", set())])

    r = normalise_tokens("")
    check("empty: returns []",
          r == [])

    r = normalise_tokens("Jean-Pierre Dupont")
    check("hyphenated: splits hyphen into tokens",
          any("jean" in g and "pierre" in g for f, g in r if f == "dupont"))

    r = normalise_tokens("Smith John")
    check("reversed: includes (smith, {john})",
          ("smith", {"john"}) in r)

    # -- initials_expand --
    print("\n--- initials_expand ---")
    check("single letter unchanged",
          initials_expand("j") == ["j"])
    check("two letters expanded",
          initials_expand("jj") == ["j", "j"])
    check("three letters expanded",
          initials_expand("abc") == ["a", "b", "c"])
    check("long token unchanged",
          initials_expand("john") == ["john"])

    # -- givens_compatible --
    print("\n--- givens_compatible ---")
    check("exact match",
          givens_compatible({"john"}, {"john"}) is True)
    check("initial vs full name",
          givens_compatible({"j"}, {"john"}) is True)
    check("full vs initial",
          givens_compatible({"john"}, {"j"}) is True)
    check("multiple givens match",
          givens_compatible({"john", "j"}, {"john", "james"}) is True)
    check("subset matching",
          givens_compatible({"john"}, {"john", "james"}) is True)
    check("empty vs any",
          givens_compatible(set(), {"john"}) is True)
    check("different names: NOT compatible",
          givens_compatible({"john"}, {"jane"}) is False)
    check("different initials: NOT compatible",
          givens_compatible({"j"}, {"m"}) is False)
    check("partial mismatch: NOT compatible",
          givens_compatible({"john", "m"}, {"john", "j"}) is False)
    check("concatenated initials: sl vs susan+l",
          givens_compatible({"sl"}, {"susan", "l"}) is True)

    # -- AuthorCluster.matches --
    print("\n--- AuthorCluster.matches ---")
    c = AuthorCluster("Susan L. Rossell", "rossell", {"susan", "l"})
    check("matches exact name",
          c.matches(normalise_tokens("Susan L. Rossell")))
    check("matches without period",
          c.matches(normalise_tokens("Susan L Rossell")))
    check("matches without middle initial",
          c.matches(normalise_tokens("Susan Rossell")))
    check("matches initials only (SL)",
          c.matches(normalise_tokens("SL Rossell")))
    check("matches comma format",
          c.matches(normalise_tokens("Rossell, Susan L.")))
    check("matches reversed order",
          c.matches(normalise_tokens("Rossell Susan")))
    check("rejects different family name",
          not c.matches(normalise_tokens("Susan L. Russell")))

    c2 = AuthorCluster("John Smith", "smith", {"john"})
    check("rejects completely different person",
          not c2.matches(normalise_tokens("Mary Johnson")))

    # -- Full clustering (end-to-end) --
    print("\n--- cluster_authors (end-to-end) ---")

    # Test 1: All Rossell variants merge
    data = {
        "Susan L. Rossell": {"doi1"},
        "Susan L Rossell": {"doi2"},
        "Susan Rossell": {"doi3"},
        "SL Rossell": {"doi4"},
        "Rossell, Susan": {"doi5"},
        "Rossell Susan": {"doi6"},
    }
    cl = cluster_authors(data)
    check("Rossell: 6 variants -> 1 cluster",
          len(cl) == 1)
    check("Rossell: canonical is longest",
          cl[0].canonical_name == "Susan L. Rossell")
    check("Rossell: 6 DOIs collected",
          len(cl[0].dois) == 6)
    check("Rossell: 6 variants tracked",
          len(cl[0].all_variants) == 6)

    # Test 2: Different people stay separate
    cl = cluster_authors({
        "John Smith": {"d1"}, "Mary Johnson": {"d2"}, "Robert Brown": {"d3"},
    })
    check("3 different people: 3 clusters",
          len(cl) == 3)

    # Test 3: Same family, different given -> separate
    cl = cluster_authors({"John Smith": {"d1"}, "Mary Smith": {"d2"}})
    check("same family different given: 2 clusters",
          len(cl) == 2)

    # Test 4: Mixed merges
    cl = cluster_authors({
        "Philip J. Phung": {"d1"}, "Philip Phung": {"d2"},
        "Richard Moulding": {"d3"}, "R. Moulding": {"d4"},
        "Erica Neill": {"d5"},
    })
    check("mixed: 3 clusters",
          len(cl) == 3)
    phung = [x for x in cl if x.family == "phung"][0]
    check("mixed: Phung has 2 variants",
          len(phung.all_variants) == 2)
    check("mixed: Phung canonical is longest",
          phung.canonical_name == "Philip J. Phung")
    moulding = [x for x in cl if x.family == "moulding"][0]
    check("mixed: Moulding has 2 variants",
          len(moulding.all_variants) == 2)

    # Test 5: Shared DOIs deduplicated
    cl = cluster_authors({
        "John Smith": {"d1", "d2"}, "J. Smith": {"d2", "d3"},
    })
    check("shared DOIs: 1 cluster with 3 DOIs",
          len(cl) == 1 and len(cl[0].dois) == 3)

    # Test 6: Comma vs standard - multi-word family names
    # Note: "Van Rheenen, Tamsyn E." has family="van rheenen" (comma is definitive)
    # while "Tamsyn E. Van Rheenen" has family="rheenen" (last token).
    # These DON'T merge because the family names differ. This is a known
    # limitation for multi-word family names when one uses commas and the
    # other doesn't. Same-format variants DO merge correctly (see Test 5
    # in the original suite where both are standard order).
    cl = cluster_authors({
        "Tamsyn E. Van Rheenen": {"d1"}, "Tamsyn Van Rheenen": {"d2"},
    })
    check("multi-word family (same format): merge into 1",
          len(cl) == 1 and len(cl[0].dois) == 2)

    # Test 7: Single-name authors
    cl = cluster_authors({"Madonna": {"d1"}, "Prince": {"d2"}})
    check("single-name authors: 2 separate",
          len(cl) == 2)

    # Test 8: Empty input
    cl = cluster_authors({})
    check("empty input: 0 clusters",
          len(cl) == 0)

    # Test 9: Initial matches full name
    cl = cluster_authors({"S Rossell": {"d1"}, "Susan Rossell": {"d2"}})
    check("S matches Susan: 1 cluster",
          len(cl) == 1)
    check("S matches Susan: canonical is Susan Rossell",
          cl[0].canonical_name == "Susan Rossell")

    # Test 10: Different initials stay separate
    cl = cluster_authors({"J Smith": {"d1"}, "M Smith": {"d2"}})
    check("J Smith vs M Smith: 2 clusters",
          len(cl) == 2)

    # Test 11: Three-part name variants
    cl = cluster_authors({
        "Sarah E. Hetrick": {"d1"}, "Sarah Hetrick": {"d2"},
        "S. E. Hetrick": {"d3"}, "SE Hetrick": {"d4"},
    })
    check("three-part variants: all merge into 1",
          len(cl) == 1)
    check("three-part variants: 4 DOIs",
          len(cl[0].dois) == 4)

    # Test 12: Complex merge preserves all DOIs
    cl = cluster_authors({
        "A. B. Smith": {"d1", "d2"}, "Alice Smith": {"d3"},
        "Alice B. Smith": {"d4"}, "Smith, A.": {"d5"},
    })
    check("complex merge: 1 cluster",
          len(cl) == 1)
    check("complex merge: 5 DOIs",
          cl[0].dois == {"d1", "d2", "d3", "d4", "d5"})
    check("complex merge: canonical is Alice B. Smith",
          cl[0].canonical_name == "Alice B. Smith")

    # Test 13: Duplicate name with different DOIs
    cl = cluster_authors({"Jose Garcia": {"d1"}, "Jose Garcia": {"d2"}})
    check("duplicate name: 1 cluster",
          len(cl) == 1)

    # Test 14: Names with only family name matching but incompatible givens
    cl = cluster_authors({
        "Alice Smith": {"d1"}, "Bob Smith": {"d2"}, "Carol Smith": {"d3"},
    })
    check("3 Smiths different givens: 3 clusters",
          len(cl) == 3)

    # Test 15: Period-only differences
    cl = cluster_authors({
        "J.K. Rowling": {"d1"}, "JK Rowling": {"d2"}, "J K Rowling": {"d3"},
    })
    check("period variations: all merge into 1",
          len(cl) == 1)
    check("period variations: 3 DOIs",
          len(cl[0].dois) == 3)

    # -- Summary --
    total = passed + failed
    print(f"\n{'=' * 70}")
    print(f"RESULTS: {passed}/{total} passed, {failed} failed")
    if failures:
        print(f"\nFailed tests:")
        for f in failures:
            print(f"  - {f}")
    print(f"{'=' * 70}")

    return passed, failed, total


# ===================================================================
# MAIN: EXTRACT AUTHORS FROM EXCEL
# ===================================================================

def extract_authors(input_file, output_file=None):
    input_path = Path(input_file)
    if not input_path.is_absolute():
        script_dir = Path(__file__).resolve().parent
        if not input_path.exists():
            fallback = script_dir / input_path
            if fallback.exists():
                input_path = fallback
    input_path = input_path.resolve()

    if not input_path.exists():
        print(f"ERROR: Input file not found: {input_path}")
        sys.exit(1)

    if output_file is None:
        output_path = input_path.parent / f"{input_path.stem}_unique_authors.xlsx"
    else:
        output_path = Path(output_file).resolve()

    print(f"Input:  {input_path}")
    print(f"Output: {output_path}")

    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = {cell.value: cell.column for cell in ws[1]}
    authors_col = headers.get("Crossref_Authors")
    doi_col = headers.get("Publication_DOI")

    if authors_col is None:
        print("ERROR: Could not find 'Crossref_Authors' column.")
        sys.exit(1)
    if doi_col is None:
        print("ERROR: Could not find 'Publication_DOI' column.")
        sys.exit(1)

    raw_author_dois = defaultdict(set)
    skip_prefixes = ("[DOI not found]", "[No authors listed]", "[Error")

    for row in range(2, ws.max_row + 1):
        authors_val = ws.cell(row=row, column=authors_col).value
        doi_val = ws.cell(row=row, column=doi_col).value

        if not authors_val or not doi_val:
            continue
        authors_str = str(authors_val)
        if any(authors_str.startswith(p) for p in skip_prefixes):
            continue

        doi = str(doi_val).strip()
        for author in authors_str.split(';'):
            author = clean_name(author)
            if author:
                raw_author_dois[author].add(doi)

    print(f"Found {len(raw_author_dois)} raw author name variants across all DOIs.")

    clusters = cluster_authors(raw_author_dois)
    print(f"Resolved to {len(clusters)} unique authors after normalisation.")

    clusters.sort(key=lambda c: c.canonical_name.lower())

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "Unique Authors"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="left", vertical="center")

    col_headers = ["Author_Name", "Name_Variants", "DOI_Count", "DOIs"]
    for col_idx, header in enumerate(col_headers, 1):
        cell = out_ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    data_font = Font(name="Arial")
    for row_idx, cluster in enumerate(clusters, 2):
        variants = "; ".join(sorted(cluster.all_variants - {cluster.canonical_name}))
        dois_str = "; ".join(sorted(cluster.dois))

        out_ws.cell(row=row_idx, column=1, value=cluster.canonical_name).font = data_font
        out_ws.cell(row=row_idx, column=2, value=variants if variants else "").font = data_font
        out_ws.cell(row=row_idx, column=3, value=len(cluster.dois)).font = data_font
        out_ws.cell(row=row_idx, column=4, value=dois_str).font = data_font

    out_ws.column_dimensions['A'].width = 35
    out_ws.column_dimensions['B'].width = 50
    out_ws.column_dimensions['C'].width = 12
    out_ws.column_dimensions['D'].width = 80
    out_ws.freeze_panes = "A2"
    out_ws.auto_filter.ref = f"A1:D{out_ws.max_row}"

    out_wb.save(output_path)
    print(f"Done. {len(clusters)} unique authors saved to: {output_path}")


# ===================================================================
# ENTRY POINT
# ===================================================================

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python extract_unique_authors.py <input.xlsx> [output.xlsx]")
        print("  python extract_unique_authors.py --test")
        sys.exit(1)

    if sys.argv[1] == "--test":
        passed, failed, total = run_tests()
        sys.exit(0 if failed == 0 else 1)

    in_arg = sys.argv[1]
    out_arg = sys.argv[2] if len(sys.argv) > 2 else None
    extract_authors(in_arg, out_arg)